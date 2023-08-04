import openpyxl
import json

subject=[]
chapter=[]
topic=[]


Excel=openpyxl.load_workbook("Excel_data.xlsx")


sheets_name=Excel.sheetnames

for i in range(len(sheets_name)):
    subject.append({'id':i+1,'name':sheets_name[i]})
# print(sheets_name)
# print(subject)

count = 1
topid_id=1
chapter_id=1
len1=0
for j in range(len(sheets_name)):
    current_sheet=Excel[sheets_name[j]]
    first_row_values = [cell.value for cell in current_sheet[1]]
    b = [{'id': count+number, 'chapter_name': i,'subject_id':j+1} for number,i in enumerate(first_row_values,start=0)]
    chapter.extend(b)
    count += len(first_row_values) 

    
    for col in current_sheet.iter_cols(values_only=True):
        new_col=[i for i in col[1:] if i is not None]
        len1=len1+len(new_col)
        c=[{'id':topid_id+id,'topic_name':new_col[i],'chapter_id':chapter_id} for id,i in enumerate(range(len(new_col)))]
        topid_id += len(c)
        topic.extend(c)
        chapter_id=chapter_id+1


# print(first_row_values)
# print(current_sheet)
# print(chapter)
print(topic)
# topic_pretty=json.dumps(topic,indent=4)
# print(len1)
# print(topic_pretty)

